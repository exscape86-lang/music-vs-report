import csv
import html
import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


OUT = Path(r"C:\Users\wizsr\music_data_tophyun_kkamananggyong")
DL = Path(r"C:\Users\wizsr\Downloads")
RAW = OUT / "raw"

META = {
    "title": "까만안경",
    "artist": "탑현",
    "release_at_kst": "2026-04-16 18:00",
    "melon_song_id": "601786200",
    "melon_album_id": "13331510",
    "genie_song_id": "114708978",
    "melon_song_url": "https://www.melon.com/song/detail.htm?songId=601786200",
    "melon_album_url": "https://www.melon.com/album/detail.htm?albumId=13331510",
    "genie_song_url": "https://www.genie.co.kr/detail/songInfo?xgnm=114708978",
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/125.0 Safari/537.36",
    "Referer": "https://www.melon.com/",
}


def now_kst() -> datetime:
    # The host session is configured for Asia/Seoul.
    return datetime.now().astimezone()


def ts() -> str:
    return now_kst().strftime("%Y%m%d_%H%M%S")


def ensure_dirs() -> None:
    for p in [
        OUT,
        RAW / "melon",
        RAW / "genie",
        RAW / "guyso",
        RAW / "youtube",
    ]:
        p.mkdir(parents=True, exist_ok=True)


def fetch(url: str, raw_dir: Path, label: str) -> dict:
    captured = now_kst()
    stamp = captured.strftime("%Y%m%d_%H%M%S")
    item = {
        "captured_at_kst": captured.isoformat(),
        "label": label,
        "url": url,
        "ok": False,
        "status_code": None,
        "content_type": None,
        "raw_path": None,
        "error": None,
    }
    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.encoding = r.encoding or "utf-8"
        suffix = ".json" if "json" in (r.headers.get("content-type") or "") else ".html"
        path = raw_dir / f"{stamp}_{label}{suffix}"
        path.write_text(r.text, encoding="utf-8", errors="replace")
        item.update(
            {
                "ok": 200 <= r.status_code < 400,
                "status_code": r.status_code,
                "content_type": r.headers.get("content-type"),
                "raw_path": str(path),
                "text": r.text,
            }
        )
    except Exception as exc:
        item["error"] = repr(exc)
    return item


def next_data(text: str):
    m = re.search(r'<script id="__NEXT_DATA__" type="application/json">(.*?)</script>', text)
    if not m:
        return None
    try:
        return json.loads(html.unescape(m.group(1)))
    except Exception:
        return None


def parse_guyso_next(fetch_item: dict) -> dict:
    parsed = {
        "label": fetch_item["label"],
        "status_code": fetch_item.get("status_code"),
        "chart": None,
        "time_unit": None,
        "columns": None,
        "rows": [],
        "song": None,
        "raw_path": fetch_item.get("raw_path"),
    }
    data = next_data(fetch_item.get("text") or "")
    if not data:
        return parsed
    page = data.get("props", {}).get("pageProps", {})
    chart_data = page.get("data") or {}
    parsed["song"] = page.get("song")
    parsed["chart"] = chart_data.get("chart")
    parsed["time_unit"] = chart_data.get("timeUnit")
    parsed["columns"] = chart_data.get("columnName")
    parsed["rows"] = chart_data.get("data") or []
    return parsed


def parse_melon_detail(fetch_item: dict) -> dict:
    text = fetch_item.get("text") or ""
    clean = html.unescape(re.sub(r"<[^>]+>", " ", text))
    return {
        "status_code": fetch_item.get("status_code"),
        "title_found": "이 사랑" in clean,
        "artist_found": "우디" in clean or "Woody" in clean,
        "album_id_candidates": sorted(set(re.findall(r"albumId=(\d+)", text))),
        "raw_path": fetch_item.get("raw_path"),
    }


def append_snapshot(summary: dict) -> None:
    path = OUT / "snapshots_hourly.csv"
    fields = [
        "captured_at_kst",
        "melon_song_id",
        "melon_album_id",
        "genie_song_id",
        "melon_detail_status",
        "melon_daily_status",
        "melon_daily_rows",
        "melon_streaming_status",
        "melon_streaming_rows",
        "genie_realtime_status",
        "genie_realtime_rows",
        "genie_latest_time",
        "genie_latest_rank",
        "genie_latest_listeners",
        "genie_latest_plays",
        "genie_latest_repeat_multiplier",
        "availability_note",
    ]
    exists = path.exists()
    with path.open("a", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        if not exists:
            w.writeheader()
        w.writerow({k: summary.get(k) for k in fields})


def write_rows_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        if not path.exists():
            path.write_text("", encoding="utf-8-sig")
        return
    keys: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in keys:
                keys.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v in (None, "", "None"):
        return None
    try:
        return float(str(v).replace(",", ""))
    except Exception:
        return None


def fmt(v) -> str:
    n = fnum(v)
    return "-" if n is None else f"{int(n):,}"


def ratio(v) -> str:
    n = fnum(v)
    return "-" if n is None else f"{n:.2f}배"


def band(v) -> str:
    n = fnum(v)
    if n is None:
        return "muted"
    if n >= 2.6:
        return "hot"
    if n >= 2.0:
        return "high"
    if n >= 1.8:
        return "mid"
    return "low"


def enrich_genie_rows(rows: list[dict]) -> list[dict]:
    out = []
    for r in rows:
        row = dict(r)
        c1 = fnum(row.get("count1"))
        c2 = fnum(row.get("count2"))
        d1 = fnum(row.get("count1Diff"))
        d2 = fnum(row.get("count2Diff"))
        row["cumulative_repeat_multiplier"] = c2 / c1 if c1 else None
        row["hourly_repeat_multiplier"] = d2 / d1 if d1 else None
        out.append(row)
    return out


def make_report(summary: dict, genie_rows: list[dict], availability: list[dict]) -> dict:
    latest = genie_rows[-1] if genie_rows else {}
    release_rows = [
        r for r in genie_rows if str(r.get("time") or "").startswith("2026-05-24")
    ]
    peak = None
    if release_rows:
        peak = max(release_rows, key=lambda r: fnum(r.get("cumulative_repeat_multiplier")) or -1)

    css = """
<style>
body{margin:0;background:#080b12;color:#f5f7fb;font:14px/1.62 -apple-system,BlinkMacSystemFont,'Segoe UI','Malgun Gothic',Arial,sans-serif}
.wrap{max-width:1380px;margin:auto;padding:26px}.hero,.card{background:#111827;border:1px solid #2b3b55;border-radius:22px;padding:20px;margin:16px 0;box-shadow:0 14px 40px #0006}
h1{margin:0 0 8px;font-size:34px;letter-spacing:-.045em}h2{margin:0 0 12px;font-size:22px}.muted{color:#aab7c8}.grid{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}
.value{font-size:27px;font-weight:850}.hot{color:#ff5d73}.high{color:#ff9f43}.mid{color:#ffd166}.low{color:#55e6a5}.mutedv{color:#aab7c8}
.note{border-left:4px solid #ffd166;background:#ffd16618;border-radius:12px;padding:12px 14px}.tablebox{overflow:auto;border:1px solid #2b3b55;border-radius:16px;max-height:680px}
table{border-collapse:separate;border-spacing:0;width:100%;min-width:980px}th,td{padding:8px 10px;border-bottom:1px solid #2b3b55;text-align:right;white-space:nowrap}
th{position:sticky;top:0;background:#182235;color:#d9e8ff}td:first-child,th:first-child{text-align:left;position:sticky;left:0;background:#111827}
.pill{display:inline-block;padding:3px 9px;border-radius:999px;background:#20304d;color:#d8e7ff;font-weight:800}.pill.hot{background:#4a1b25;color:#ff9caf}.pill.high{background:#472c16;color:#ffbd7a}.pill.mid{background:#463b18;color:#ffe08a}.pill.low{background:#173927;color:#91f3bf}
@media(max-width:900px){.grid{grid-template-columns:1fr}.wrap{padding:12px}}
</style>
"""
    parts = [
        "<!doctype html><html lang='ko'><head><meta charset='utf-8'>"
        "<meta name='viewport' content='width=device-width,initial-scale=1'>"
        f"<title>{META['artist']} - {META['title']} 시간별 반복재생 배수</title>{css}</head><body><div class='wrap'>",
        "<section class='hero'>"
        f"<h1>{META['artist']} - {META['title']}</h1>"
        f"<div class='muted'>발매: {META['release_at_kst']} KST | "
        f"멜론 songId {META['melon_song_id']} / albumId {META['melon_album_id']} | "
        f"지니 songId {META['genie_song_id']}</div>"
        "<p class='note'><b>반복재생 배수</b> = 누적 재생 수 ÷ 누적 청취자 수. "
        "시간별 배수는 시간별 재생 증가분 ÷ 시간별 청취자 증가분입니다. "
        "1.8배 이상, 2.0배 이상, 2.6배 이상 구간을 강조합니다.</p></section>",
        "<section class='grid'>",
        f"<div class='card'><div class='muted'>최신 수집 시각</div><div class='value'>{summary.get('captured_at_kst','-')}</div></div>",
        f"<div class='card'><div class='muted'>지니 최신 누적 배수</div><div class='value {band(latest.get('cumulative_repeat_multiplier'))}'>{ratio(latest.get('cumulative_repeat_multiplier'))}</div></div>",
        f"<div class='card'><div class='muted'>지니 최신 청취자/재생</div><div class='value'>{fmt(latest.get('count1'))} / {fmt(latest.get('count2'))}</div></div>",
        f"<div class='card'><div class='muted'>발매 당일 최고 배수</div><div class='value {band((peak or {}).get('cumulative_repeat_multiplier'))}'>{ratio((peak or {}).get('cumulative_repeat_multiplier'))}</div></div>",
        "</section>",
        "<section class='card'><h2>수집 가능성 요약</h2><div class='tablebox'><table><thead><tr>"
        "<th>소스</th><th>HTTP</th><th>차트</th><th>단위</th><th>행 수</th><th>비고</th></tr></thead><tbody>",
    ]
    for a in availability:
        parts.append(
            f"<tr><td>{html.escape(a.get('label',''))}</td><td>{a.get('status_code')}</td>"
            f"<td>{html.escape(str(a.get('chart') or '-'))}</td><td>{html.escape(str(a.get('time_unit') or '-'))}</td>"
            f"<td>{len(a.get('rows') or [])}</td><td>{html.escape(a.get('note') or '')}</td></tr>"
        )
    parts.append("</tbody></table></div></section>")

    parts.append(
        "<section class='card'><h2>지니 시간별 반복재생 배수</h2><div class='tablebox'><table><thead><tr>"
        "<th>시간</th><th>순위</th><th>누적 청취자</th><th>누적 재생</th><th>누적 반복재생 배수</th>"
        "<th>시간당 청취자 증가</th><th>시간당 재생 증가</th><th>시간당 반복재생 배수</th></tr></thead><tbody>"
    )
    for r in reversed(genie_rows):
        parts.append(
            f"<tr><td>{html.escape(str(r.get('time',''))).replace('T',' ')}</td>"
            f"<td>{fmt(r.get('ranking'))}</td><td>{fmt(r.get('count1'))}</td><td>{fmt(r.get('count2'))}</td>"
            f"<td><span class='pill {band(r.get('cumulative_repeat_multiplier'))}'>{ratio(r.get('cumulative_repeat_multiplier'))}</span></td>"
            f"<td>{fmt(r.get('count1Diff'))}</td><td>{fmt(r.get('count2Diff'))}</td>"
            f"<td><span class='pill {band(r.get('hourly_repeat_multiplier'))}'>{ratio(r.get('hourly_repeat_multiplier'))}</span></td></tr>"
        )
    parts.append("</tbody></table></div></section>")
    parts.append(
        "<section class='card'><h2>파일</h2><ul>"
        f"<li>작업 폴더: {OUT}</li>"
        "<li>raw 원본은 플랫폼/시각별로 보존됩니다.</li>"
        "<li>멜론 발매 직후 일별/스트리밍 카드 데이터는 아직 공개되지 않을 수 있습니다.</li>"
        "</ul></section></div></body></html>"
    )
    report = OUT / "woody_isarang_hourly_repeat_multiplier_report.html"
    report.write_text("\n".join(parts), encoding="utf-8")
    dl_report = DL / "woody_isarang_hourly_repeat_multiplier_report.html"
    shutil.copy2(report, dl_report)
    text = report.read_text(encoding="utf-8")
    qc = {
        "report": str(report),
        "download": str(dl_report),
        "replacement_char": "\ufffd" in text,
        "title_present": META["title"] in text,
        "artist_present": "우디" in text and "Woody" in text,
        "formula_present": "누적 재생 수 ÷ 누적 청취자 수" in text,
        "rows": len(genie_rows),
    }
    qc["pass"] = (
        not qc["replacement_char"]
        and qc["title_present"]
        and qc["artist_present"]
        and qc["formula_present"]
    )
    (OUT / "qc_result_hourly.json").write_text(json.dumps(qc, ensure_ascii=False, indent=2), encoding="utf-8")
    return qc


def make_xlsx(genie_rows: list[dict], availability: list[dict]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "요약"
    ws.append(["항목", "값"])
    for k, v in META.items():
        ws.append([k, v])
    ws.append(["생성시각", now_kst().isoformat()])
    ws.append(["반복재생 배수", "누적 재생 수 / 누적 청취자 수"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws2 = wb.create_sheet("지니_시간별")
    if genie_rows:
        headers = list(genie_rows[0].keys())
        ws2.append(headers)
        for row in genie_rows:
            ws2.append([row.get(h) for h in headers])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E8FF")

    ws3 = wb.create_sheet("수집가능성")
    headers = ["label", "url", "status_code", "chart", "time_unit", "columns", "row_count", "note"]
    ws3.append(headers)
    for a in availability:
        ws3.append(
            [
                a.get("label"),
                a.get("url"),
                a.get("status_code"),
                a.get("chart"),
                a.get("time_unit"),
                a.get("columns"),
                len(a.get("rows") or []),
                a.get("note"),
            ]
        )
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E8FF")

    csv_path = OUT / "snapshots_hourly.csv"
    ws4 = wb.create_sheet("스냅샷_로그")
    if csv_path.exists() and csv_path.read_text(encoding="utf-8-sig").strip():
        with csv_path.open(encoding="utf-8-sig", newline="") as f:
            for row in csv.reader(f):
                ws4.append(row)
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 70)
    path = OUT / "woody_isarang_data_pack.xlsx"
    wb.save(path)
    shutil.copy2(path, DL / path.name)
    return str(path)


def run_once() -> dict:
    ensure_dirs()
    stamp = ts()
    urls = {
        "melon_detail": META["melon_song_url"],
        "melon_album": META["melon_album_url"],
        "melon_daily_users": f"https://xn--o39an51b2re.com/chart/melon/daily/trend/graph?songIds={META['melon_song_id']}",
        "melon_streaming_card_daily": f"https://xn--o39an51b2re.com/chart/melon/streaming-card/trend/graph?songIds={META['melon_song_id']}",
        "genie_realtime_hourly": f"https://xn--o39an51b2re.com/chart/genie/realtime/trend/ranking/{META['genie_song_id']}",
        "genie_daily_rank": f"https://xn--o39an51b2re.com/chart/genie/daily/trend/ranking/{META['genie_song_id']}",
        "genie_detail": META["genie_song_url"],
        "melon_search": "https://www.melon.com/search/total/index.htm?q=" + quote("우디 이 사랑"),
        "genie_search": "https://www.genie.co.kr/search/searchMain?query=" + quote("우디 이 사랑"),
    }
    fetched = {}
    for label, url in urls.items():
        if label.startswith("melon"):
            raw_dir = RAW / "melon" if "guyso" not in label else RAW / "guyso"
        elif label.startswith("genie"):
            raw_dir = RAW / "genie"
        else:
            raw_dir = RAW / "guyso"
        if "daily" in label or "streaming" in label or "realtime" in label:
            raw_dir = RAW / "guyso"
        fetched[label] = fetch(url, raw_dir, label)

    parsed_melon = parse_melon_detail(fetched["melon_detail"])
    guyso_labels = [
        "melon_daily_users",
        "melon_streaming_card_daily",
        "genie_realtime_hourly",
        "genie_daily_rank",
    ]
    parsed = {label: parse_guyso_next(fetched[label]) for label in guyso_labels}
    genie_rows = enrich_genie_rows(parsed["genie_realtime_hourly"]["rows"])

    write_rows_csv(OUT / "genie_realtime_hourly_listener_play.csv", genie_rows)
    write_rows_csv(OUT / "melon_daily_users.csv", parsed["melon_daily_users"]["rows"])
    write_rows_csv(OUT / "melon_streaming_card_daily.csv", parsed["melon_streaming_card_daily"]["rows"])

    availability = []
    for label in guyso_labels:
        p = parsed[label]
        note = ""
        if not p["rows"]:
            note = "아직 공개 데이터 없음 또는 GuySome 500/빈 데이터"
        availability.append(
            {
                "label": label,
                "url": urls[label],
                "status_code": fetched[label].get("status_code"),
                "chart": p.get("chart"),
                "time_unit": p.get("time_unit"),
                "columns": json.dumps(p.get("columns"), ensure_ascii=False),
                "rows": p.get("rows") or [],
                "note": note,
            }
        )
    with (OUT / "availability_checks.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(
            f, fieldnames=["label", "url", "status_code", "chart", "time_unit", "columns", "rows", "note"]
        )
        w.writeheader()
        for a in availability:
            row = dict(a)
            row["rows"] = len(a["rows"])
            w.writerow(row)

    latest = genie_rows[-1] if genie_rows else {}
    summary = {
        "captured_at_kst": now_kst().isoformat(),
        "melon_song_id": META["melon_song_id"],
        "melon_album_id": META["melon_album_id"],
        "genie_song_id": META["genie_song_id"],
        "melon_detail_status": fetched["melon_detail"].get("status_code"),
        "melon_daily_status": fetched["melon_daily_users"].get("status_code"),
        "melon_daily_rows": len(parsed["melon_daily_users"]["rows"]),
        "melon_streaming_status": fetched["melon_streaming_card_daily"].get("status_code"),
        "melon_streaming_rows": len(parsed["melon_streaming_card_daily"]["rows"]),
        "genie_realtime_status": fetched["genie_realtime_hourly"].get("status_code"),
        "genie_realtime_rows": len(genie_rows),
        "genie_latest_time": latest.get("time"),
        "genie_latest_rank": latest.get("ranking"),
        "genie_latest_listeners": latest.get("count1"),
        "genie_latest_plays": latest.get("count2"),
        "genie_latest_repeat_multiplier": latest.get("cumulative_repeat_multiplier"),
        "availability_note": "first hourly collector run",
    }
    append_snapshot(summary)

    payload = {
        "meta": META,
        "stamp": stamp,
        "melon_detail": parsed_melon,
        "availability": [
            {k: v for k, v in a.items() if k != "rows"} | {"row_count": len(a["rows"])}
            for a in availability
        ],
        "summary": summary,
    }
    (OUT / f"snapshot_{stamp}.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT / "latest_snapshot.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    qc = make_report(summary, genie_rows, availability)
    xlsx = make_xlsx(genie_rows, availability)
    payload["qc"] = qc
    payload["xlsx"] = xlsx
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return payload


if __name__ == "__main__":
    run_once()
